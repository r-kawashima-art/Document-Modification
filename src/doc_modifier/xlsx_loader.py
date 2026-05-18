"""Load applicant rows from an Excel data file.

The file's first row is treated as the header (見出し). All subsequent
non-empty rows are yielded as ``dict[str, str]`` keyed by the header.
Header names are normalized to ``snake_case`` lower so users can write
``Passport No.`` or ``passport_no`` interchangeably.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


_NORMALIZE_RE = re.compile(r"[^0-9a-z]+")


def normalize_header(text: str) -> str:
    """Normalize ``'Passport No.'`` → ``'passport_no'``.

    Trailing/leading underscores are stripped. Empty result becomes ''.
    """
    if text is None:
        return ""
    norm = _NORMALIZE_RE.sub("_", str(text).strip().lower()).strip("_")
    return norm


def _format_cell(value):
    """Render an openpyxl cell value as the string that will go into the doc."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Default to dd/mm/yyyy to match the original template (25/07/1969).
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def load_rows(path: str | Path, sheet: str | None = None) -> list[dict[str, str]]:
    """Load all data rows as a list of dicts.

    :param path: Path to the ``.xlsx`` data file.
    :param sheet: Optional sheet name. Defaults to the active sheet.
    :returns: One dict per non-empty data row.
    """
    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows: Iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = [normalize_header(h) for h in header_row]
    out: list[dict[str, str]] = []
    for raw in rows:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        row: dict[str, str] = {}
        for h, v in zip(headers, raw):
            if not h:
                continue
            row[h] = _format_cell(v)
        out.append(row)
    return out
