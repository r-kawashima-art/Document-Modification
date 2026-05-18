"""``{{token}}`` replacement inside ``.xlsx`` templates.

Cell-level formatting is stored separately from cell *values* in
openpyxl, so a simple ``cell.value = cell.value.replace(...)`` keeps
fonts, borders, fills, number formats, etc. fully intact.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook


TOKEN_RE = re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}")


def _substitute(text: str, mapping: Mapping[str, str]) -> tuple[str, int]:
    count = 0

    def _repl(m):
        nonlocal count
        key = m.group(1)
        if key in mapping:
            count += 1
            return str(mapping[key])
        return m.group(0)

    return TOKEN_RE.sub(_repl, text), count


def render(template_path: str | Path, mapping: Mapping[str, str], out_path: str | Path) -> int:
    """Render an Excel template, writing to ``out_path``.

    :returns: total number of token substitutions performed.
    """
    wb = load_workbook(filename=str(template_path))
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    new_val, n = _substitute(cell.value, mapping)
                    if n:
                        cell.value = new_val
                        total += n
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return total


def find_tokens(template_path: str | Path) -> set[str]:
    wb = load_workbook(filename=str(template_path))
    found: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for m in TOKEN_RE.finditer(cell.value):
                        found.add(m.group(1))
    return found
